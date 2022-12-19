#script that reads a list of plate-data and when given parameters, will output scores for antibody screening
#uses pyexcel-xlsx, download at https://pypi.org/project/pyexcel-xlsx/ or pip install pyexcel-xlsx
import numpy as np
from optparse import OptionParser
from pyexcel_xlsx import save_data
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from ast import literal_eval



parser = OptionParser()
parser.add_option("-d", "--data", dest="plate_path")
parser.add_option("-x", "--excel", dest="excel_path", default=False)
parser.add_option("-l", "--labels", dest="label_path", default=False)
parser.add_option("-c", "--conc", dest="conc_path", default=None)
parser.add_option("-t", "--thresh", dest="thresh_path", default = None)
parser.add_option("-s", "--score", dest="score_path", default = None)
parser.add_option("-f", "--format", dest="format_path", default = None )
parser.add_option("-i", "--id_con", dest="id_con_path",default = None)

opts,args = parser.parse_args()

plate_file = opts.plate_path
plate_data = np.zeros((8,12))
plate_file = plate_file.replace("\ "," ")
plate_file = plate_file.replace("  "," ")
plate = open(plate_file,"r") 
plate.readline()
header = plate.readline().split()
plate.readline()
con_l = literal_eval(opts.conc_path)
format_type = opts.format_path
thresh = literal_eval(opts.thresh_path)
scor_vals = literal_eval(opts.score_path)
id_con = opts.id_con_path
print(con_l)
for i in con_l:
    print(i)

h_met = opts.format_path + ", "+ opts.conc_path + ", " + opts.label_path + ", " + opts.id_con_path  + ", " + opts.thresh_path + ", " + opts.score_path 

if len(con_l) == 3:
    form = 'p2'
else:
    form = 'p1'


def row_ave(row):
    row_out = []
    for i in range(0,11):
        if i%2 == 0:
            row_out.append((row[i]+row[i+1])/2)
    return(row_out)


def row_sen(AVG):
    sen_out = []
    for i in range(0,8):
        if i//4 == 0:
            sen_out.append(AVG[i]-AVG[3])
        else:
            sen_out.append(AVG[i]-AVG[7])
    return(sen_out)

def row_s2n(SEN):
    sen_out = []
    for i in range(0,8):
        if i//4 == 0:
            sen_out.append(AVG[i]/AVG[3])
        else:
            sen_out.append(AVG[i]/AVG[7])
    return(sen_out)

def row_intf(AVG):
    intf_out = []
    for i in range(0,4):
        intf_out.append(100*(AVG[i]-AVG[i+4])/AVG[i])
    return(intf_out)

AVG=np.zeros((8,6))
for i in range(0,8):
    plate_data[i] = plate.readline().split()[1:]
    AVG[i]=row_ave(plate_data[i])

SEN = row_sen(AVG)
SEN = np.array(SEN)

S2N = row_s2n(SEN)
S2N = np.array(S2N)

if form == 'p1':
    INTF = row_intf(AVG)
    INTF = np.array(INTF)



analytics = dict()
for i in range(0,6):
    if form == 'p1':
        analytics[str(i)] = {"ULQ" : {}, "LLQ" : {}, "S2N" : {}, "INTF" : {},"BACK" :{}}
    else:
        analytics[str(i)] = {"ULQ" : {}, "LLQ" : {}, "S2N" : {}, "INTF" : {},"BACK" :{}}

'''
def ULOQ(sen_in):
    if sen_in < 65000:
        return(0)
    if sen_in > 100000:
        return(2)
    else:
        return(1)

def LLOQ(sen_in):
    if sen_in < 500:
        return(0)
    if sen_in > 1000:
        return(2)
    else:
        return(1)

def sig2noise(s2n_in):
    if s2n_in < 3:
        return(0)
    if s2n_in > 10:
        return(3)
    elif s2n_in > 5:
        return(2)
    else:
        return(1)

def matrix(m_in):
    if abs(m_in) > 30:
        return(-2)
    if abs(m_in) < 10:
        return(2)
    elif abs(m_in) > 20:
        return(1)
    else:
        return(0)

def Back(b_in):
    if abs(b_in) > 750:
        return(-2)
    if abs(b_in) < 200:
        return(0)
    else:
        return(-1)
'''
# Background, S2N, LLT, ULT, ant intf

def tscor(v_in,ty=None):
    ty = ty #type of scorrer
    s = None 
    for i in range(len(thresh[ty])):
        if abs(v_in) < float(thresh[ty][i]):
            if s != None:
                pass
            else:
                s = scor_vals[ty][i]
    if s == None:
        s = scor_vals[ty][-1]
    s = int(s)
    return(s)


def scorer():
    #works for the data from mk-0482 12-5-18
    scores = []
    if form == 'p1':
        for i in range(0,6):
            analytics[str(i)]['INTF']['1000 ng/ml + intf'] = tscor(INTF[0][i],4) 
            analytics[str(i)]['INTF']['100 ng/ml + intf'] = tscor(INTF[1][i],4)
            analytics[str(i)]['INTF']['10 ng/ml + intf'] = tscor(INTF[2][i],4)
            analytics[str(i)]['S2N']['10 ng/ml'] = tscor(S2N[2][i],1)
            analytics[str(i)]['S2N']['10 ng/ml + 1 ug/ml ILT3'] = tscor(S2N[-2][i],1)
            analytics[str(i)]['LLQ']['10 ng/ml'] = tscor(SEN[2][i],2)
            analytics[str(i)]['LLQ']['10 ng/ml + 1 ug/ml ILT3'] = tscor(SEN[-2][i],2)
            analytics[str(i)]['ULQ']['1000 ng/ml'] = tscor(SEN[0][i],3)
            analytics[str(i)]['ULQ']['1000 ng/ml + 1 ug/ml ILT3'] = tscor(SEN[-4][i],3)
            analytics[str(i)]['BACK']['BACK'] = tscor(AVG[3][i],0)
            s = 0
            for j in analytics[str(i)]:
                for k in analytics[str(i)][j]:
                    s += analytics[str(i)][j][k]
            analytics[str(i)]['TOTAL'] = s
    else:
        for i in range(0,6):
            analytics[str(i)]['S2N']['10 ng/ml'] = tscor(S2N[2][i],1)
            analytics[str(i)]['LLQ']['10 ng/ml'] = tscor(SEN[2][i],2)
            analytics[str(i)]['ULQ']['1000 ng/ml'] = tscor(SEN[0][i],3)
            analytics[str(i)]['BACK']['BACK'] = tscor(AVG[3][i],0)
            analytics[str(i)]['S2N']['10 ng/ml b'] = tscor(S2N[-2][i],1)
            analytics[str(i)]['LLQ']['10 ng/ml b'] = tscor(SEN[-2][i],2)
            analytics[str(i)]['ULQ']['1000 ng/ml b'] = tscor(SEN[-4][i],3)
            analytics[str(i)]['BACK']['BACK b'] = tscor(AVG[7][i],0)
            s = 0
            for j in analytics[str(i)]:
                for k in analytics[str(i)][j]:
                    if k[-1] != 'b':
                        s += analytics[str(i)][j][k]
            analytics[str(i)]['TOTAL'] = s
            s = 0
            for j in analytics[str(i)]:
                print(j)
                if j != 'TOTAL':
                    print(analytics[str(i)][j])
                    for k in analytics[str(i)][j]:
                        if k[-1] == 'b':
                            s += analytics[str(i)][j][k]
            analytics[str(i)]['TOTAL b'] = s





scorer()



def plate_2_excel(x):
    row = []
    row.append(chr(ord('@')+x+1))
    for i in plate_data[x]:
        row.append(i)
    return(row)

exl_ary = []
if type(opts.excel_path) == str:
    data = OrderedDict()
    exl_ary.append([header[0]+' '+header[1],header[2]+' '+header[3]+' '+header[4],'','','','','','','','','','','','','','','','',h_met])
    r1 = ['',1, 2, 3, 4, 5, 6, 7, 8, 9 , 10, 11, 12]
    exl_ary.append(r1)
    for i in range(0,8):
        exl_ary.append(plate_2_excel(i))
    exl_ary.append([])
    exl_ary.append([])
    labels = literal_eval(opts.label_path)
    label_1 = labels[0]
    label_2 = labels[1]
    label_3 = labels[2]
    label_4 = labels[3]
    label_5 = labels[4]
    label_6 = labels[5]
    label_7 = labels[6]
    exl_ary.append(['',
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_1),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_1),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_1),
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_2),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_2),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_2),
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_3),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_3),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_3),
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_4),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_4),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_4),
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_5),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_5),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_5),
                    "Biotin - {0} & Tag - {1} AVG".format(label_7,label_6),
                    "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_6),
                    "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_6)])
    if form == 'p1':
        con =[con_l[0],con_l[1],con_l[2],'NSB',con_l[0]+' + '+con_l[3]+" Antigen",
            con_l[1]+' + '+con_l[3]+" Antigen",con_l[2]+' + '+con_l[3]+" Antigen",'NSB + '+con_l[3]+" Antigen"]
        for i in range(0,8):
            row =[con[i]]
            for j in range(0,6):
                row.append(AVG[i][j])
                row.append(SEN[i][j])
                row.append(S2N[i][j])
            exl_ary.append(row)
        exl_ary.append([])
        exl_ary.append(['',
                        "Biotin - {0} & Tag - {1}".format(label_7,label_1),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_2),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_3),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_4),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_5),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_6)])  
        con=[con_l[0]+' + '+con_l[3]+" Antigen Interference",con_l[1]+' + '+con_l[3]+" Antigen Interference",
            con_l[2]+' + '+con_l[3]+" Antigen Interference","NSB + "+con_l[3]+" Antigen Interference"]
        for i in range(0,4):
            row =[con[i]]
            for j in range(0,6):
                row.append(INTF[i][j])
                row.append("")
                row.append("")
            exl_ary.append(row)
        exl_ary.append([])
        exl_ary.append(['',
                        "Scoring Parameter",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_1),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_2),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_3),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_4),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_5),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_6)])
        con1=["Background Scoring","ULOQ Sensitivity Scoring",
            "ULOQ Sensitivity Scoring","LLOQ Sensitivity Scoring",
            "LLOQ Sensitivity Scoring","S/N Scoring",
            "S/N Scoring","Soluble Antigen Interference Scoring",
            "Soluble Antigen Interference Scoring","Soluble Antigen Interference Scoring",
            "Soluble Antigen Interference Scoring","","Total Score"]
        con2=["Background","ULOQ - 1 ug/mL ILT3 Interference",
            "ULOQ + "+con_l[3]+" Antigen Interference","LLOQ - "+con_l[3]+" Antigen Interference",
            "LLOQ + "+con_l[3]+" Antigen Interference","S/N - "+con_l[3]+" Antigen Interference",
            "S/N + "+con_l[3]+" Antigen Interference",con_l[0]+" + "+con_l[3]+" Antigen Interference",
            con_l[1]+" + "+con_l[3]+" Antigen Interference",con_l[2]+" + "+con_l[3]+" Antigen Interference",
            "",""]
        sr1 = ["BACK","ULQ","ULQ","LLQ","LLQ","S2N","S2N","INTF","INTF","INTF"]
        sr2 = ["BACK","1000 ng/ml","1000 ng/ml + 1 ug/ml ILT3",
                '10 ng/ml','10 ng/ml + 1 ug/ml ILT3',
                '10 ng/ml','10 ng/ml + 1 ug/ml ILT3',
                '1000 ng/ml + intf','100 ng/ml + intf','10 ng/ml + intf']
        for i in range(0,10):
            row=[con1[i],con2[i]]
            for j in range(0,6):
                row.append(analytics[str(j)][sr1[i]][sr2[i]])
                row.append('')
                row.append('')
            exl_ary.append(row)
        exl_ary.append([])
        row = ['Total Score','']
        for i in range(0,6):
            row.append(analytics[str(i)]['TOTAL'])
            row.append('')
            row.append('')
        exl_ary.append(row)
    else:
        con =[con_l[0],con_l[1],con_l[2],'NSB']
        for i in range(0,4):
            row =[con[i]]
            for j in range(0,6):
                row.append(AVG[i][j])
                row.append(SEN[i][j])
                row.append(S2N[i][j])
            exl_ary.append(row)
        exl_ary.append([])
        exl_ary.append([])
        exl_ary.append(['',
                        "Scoring Parameter",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_1),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_2),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_3),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_4),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_5),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_6)])
        con1=["Background Scoring","ULOQ Sensitivity Scoring","LLOQ Sensitivity Scoring","S/N Scoring","","Total Score"]
        con2=["Background","ULOQ","LLOQ","S/N","",""]
        sr1 = ["BACK","ULQ","LLQ","S2N"]
        sr2 = ["BACK","1000 ng/ml",'10 ng/ml','10 ng/ml']
        for i in range(0,4):
            row=[con1[i],con2[i]]
            for j in range(0,6):
                row.append(analytics[str(j)][sr1[i]][sr2[i]])
                row.append('')
                row.append('')
            exl_ary.append(row)
        exl_ary.append([])
        row = ['Total Score','']
        for i in range(0,6):
            row.append(analytics[str(i)]['TOTAL'])
            row.append('')
            row.append('')
        exl_ary.append(row)
        ###
        exl_ary.append([])
        label_1 = labels[7]
        label_2 = labels[8]
        label_3 = labels[9]
        label_4 = labels[10]
        label_5 = labels[11]
        label_6 = labels[12]
        label_7 = labels[13]
        exl_ary.append(['',
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_1),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_1),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_1),
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_2),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_2),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_2),
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_3),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_3),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_3),
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_4),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_4),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_4),
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_5),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_5),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_5),
                "Biotin - {0} & Tag - {1} AVG".format(label_7,label_6),
                "Biotin - {0} & Tag - {1}  Sensitivity".format(label_7,label_6),
                "Biotin - {0} & Tag - {1}  S/N".format(label_7,label_6)])
        for i in range(0,4):
            row =[con[i]]
            for j in range(0,6):
                row.append(AVG[i+4][j])
                row.append(SEN[i+4][j])
                row.append(S2N[i+4][j])
            exl_ary.append(row)
        exl_ary.append([])
        exl_ary.append([])
        exl_ary.append(['',
                        "Scoring Parameter",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_1),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_2),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_3),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_4),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_5),
                        "",
                        "",
                        "Biotin - {0} & Tag - {1}".format(label_7,label_6)])
        con1=["Background Scoring","ULOQ Sensitivity Scoring","LLOQ Sensitivity Scoring","S/N Scoring","","Total Score"]
        con2=["Background","ULOQ","LLOQ","S/N","",""]
        sr1 = ["BACK","ULQ","LLQ","S2N"]
        sr2 = ["BACK","1000 ng/ml b",'10 ng/ml b','10 ng/ml b']
        for i in range(0,4):
            row=[con1[i],con2[i]]
            for j in range(0,6):
                row.append(analytics[str(j)][sr1[i]][sr2[i]])
                row.append('')
                row.append('')
            exl_ary.append(row)
        exl_ary.append([])
        row = ['Total Score','']
        for i in range(0,6):
            row.append(analytics[str(i)]['TOTAL b'])
            row.append('')
            row.append('')
        exl_ary.append(row)
    data.update({"Sheet 1": exl_ary})
    save_data(opts.excel_path, data)


###FORMATING###
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill, Fill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
workbook = load_workbook(opts.excel_path)
#workbook = load_workbook(filename="mk-0482 p1_2018-12-05-154224_21H16ALG51.xlsx")
sheet = workbook.active
t =  Side(border_style="thin")

sides = Border(top=None,
                        right=t,
                        bottom=None,
                        left=t)

bottom = Border(top=None,
                        right=t,
                        bottom=t,
                        left=t)

for i in range(1,20):
    sheet.column_dimensions["{0}".format(chr(ord('@')+i))].width = 13.42578125 



if form == 'p1':
    for i in range(30,41):
        sheet["C{0}".format(i)].border = sides
        sheet["F{0}".format(i)].border = sides
        sheet["I{0}".format(i)].border = sides
        sheet["L{0}".format(i)].border = sides
        sheet["O{0}".format(i)].border = sides
        sheet["R{0}".format(i)].border = sides

    sheet["C41"].border = bottom
    sheet["F41"].border = bottom
    sheet["I41"].border = bottom
    sheet["L41"].border = bottom
    sheet["O41"].border = bottom
    sheet["R41"].border = bottom


    sheet["A30"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["A31"].fill = PatternFill(patternType='solid', fgColor="ffffc000")
    sheet["A32"].fill = PatternFill(patternType='solid', fgColor="ffffc000")
    sheet["A33"].fill = PatternFill(patternType='solid', fgColor="FF00B0F0")
    sheet["A34"].fill = PatternFill(patternType='solid', fgColor="FF00B0F0")
    sheet["A35"].fill = PatternFill(patternType='solid', fgColor="FFE6B8B7")
    sheet["A36"].fill = PatternFill(patternType='solid', fgColor="FFE6B8B7")
    sheet["A37"].fill = PatternFill(patternType='solid', fgColor="FF00B050")
    sheet["A38"].fill = PatternFill(patternType='solid', fgColor="FF00B050")
    sheet["A39"].fill = PatternFill(patternType='solid', fgColor="FF00B050")


    for i in range(18,42):
        sheet["A{0}".format(i)].alignment = Alignment(wrapText=True)

    for i in range(30,40):
        sheet["B{0}".format(i)].alignment = Alignment(wrapText=True)

    for i in range(1,20):
        sheet["{0}13".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)

    for i in range(1,20):
        sheet["{0}23".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)

    for i in range(1,20):
        sheet["{0}29".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)


    sheet.row_dimensions[13].height = 60.0
    sheet.row_dimensions[18].height = 30
    sheet.row_dimensions[19].height = 30
    sheet.row_dimensions[20].height = 30
    for i in range(23,28):
        sheet.row_dimensions[i].height = 45.0
    for i in range(30,37):
        sheet.row_dimensions[i].height = 45.0
    sheet.row_dimensions[37].height = 60.0
    sheet.row_dimensions[38].height = 60.0
    sheet.row_dimensions[39].height = 60.0
    sheet.row_dimensions[41].height = 30.0


    sheet.row_dimensions[23].height = 45.0
    sheet.row_dimensions[29].height = 60
    workbook.save(opts.excel_path)
    print("done")
else:
    for i in range(21,26):
        sheet["C{0}".format(i)].border = sides
        sheet["F{0}".format(i)].border = sides
        sheet["I{0}".format(i)].border = sides
        sheet["L{0}".format(i)].border = sides
        sheet["O{0}".format(i)].border = sides
        sheet["R{0}".format(i)].border = sides
    
    for i in range(21+15,26+15):
        sheet["C{0}".format(i)].border = sides
        sheet["F{0}".format(i)].border = sides
        sheet["I{0}".format(i)].border = sides
        sheet["L{0}".format(i)].border = sides
        sheet["O{0}".format(i)].border = sides
        sheet["R{0}".format(i)].border = sides

    sheet["C26"].border = bottom
    sheet["F26"].border = bottom
    sheet["I26"].border = bottom
    sheet["L26"].border = bottom
    sheet["O26"].border = bottom
    sheet["R26"].border = bottom

    sheet["C41"].border = bottom
    sheet["F41"].border = bottom
    sheet["I41"].border = bottom
    sheet["L41"].border = bottom
    sheet["O41"].border = bottom
    sheet["R41"].border = bottom


    sheet["A21"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["A22"].fill = PatternFill(patternType='solid', fgColor="ffffc000")
    sheet["A23"].fill = PatternFill(patternType='solid', fgColor="FF00B0F0")
    sheet["A24"].fill = PatternFill(patternType='solid', fgColor="FFE6B8B7")

    sheet["A36"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["A37"].fill = PatternFill(patternType='solid', fgColor="ffffc000")
    sheet["A38"].fill = PatternFill(patternType='solid', fgColor="FF00B0F0")
    sheet["A39"].fill = PatternFill(patternType='solid', fgColor="FFE6B8B7")

    for i in range(18,42):
        sheet["A{0}".format(i)].alignment = Alignment(wrapText=True)

    for i in range(30,40):
        sheet["B{0}".format(i)].alignment = Alignment(wrapText=True)

    for i in range(1,20):
        sheet["{0}13".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)

    for i in range(1,20):
        sheet["{0}28".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)


    for i in range(1,20):
        sheet["{0}20".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)
    
    for i in range(1,20):
        sheet["{0}35".format(chr(ord('@')+i))].alignment = Alignment(wrapText=True)


    sheet.row_dimensions[13].height = 60.0
    sheet.row_dimensions[18].height = 30
    sheet.row_dimensions[13+15].height = 60.0
    sheet.row_dimensions[18+15].height = 30

    for i in range(21,25):
        sheet.row_dimensions[i].height = 45.0
    sheet.row_dimensions[26].height = 30.0

    for i in range(21+15,25+15):
        sheet.row_dimensions[i].height = 45.0
    sheet.row_dimensions[26+15].height = 30.0


    sheet.row_dimensions[20].height = 60
    sheet.row_dimensions[20+15].height = 60
    workbook.save(opts.excel_path)    
    print("done")











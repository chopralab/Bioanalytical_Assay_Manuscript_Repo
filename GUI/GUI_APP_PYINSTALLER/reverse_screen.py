#script that reads a list of reverse screen plate-data and when given parameters, and output scores
#uses pyexcel-xlsx, download at https://pypi.org/project/pyexcel-xlsx/ or pip install pyexcel-xlsx
import numpy as np
from optparse import OptionParser
from pyexcel_xlsx import save_data
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from ast import literal_eval


parser = OptionParser()
parser.add_option("-d", "--data", dest="plate_path")
parser.add_option("-x", "--excel", dest="excel_path", default=False)
parser.add_option("-l", "--labels", dest="label_path", default=False)
parser.add_option("-c", "--conc", dest="conc_path", default=None)
parser.add_option("-t", "--thresh", dest="thresh_path", default = None)
parser.add_option("-s", "--score", dest="score_path", default = None)
opts,args = parser.parse_args()

h_met = opts.conc_path + ", " + opts.label_path + ", " + opts.thresh_path + ", " + opts.score_path 



plate_file = opts.plate_path
plate_file = plate_file.replace("\ "," ")
plate_file = plate_file.replace("  "," ")
plate_data = np.zeros((8,12))
plate = open(plate_file,"r") 
plate.readline()
header = plate.readline().split()
plate.readline()
con_l = literal_eval(opts.conc_path)

if len(con_l) == 3:
    form = 'r2'
else:
    form = 'r1'

thresh = literal_eval(opts.thresh_path)
scor_vals = literal_eval(opts.score_path)


def col_intf(AVG):
    if form == 'r1':
        intf_out = []
        intf_out.append(100*(AVG[0]-AVG[1])/AVG[0])
        intf_out.append(100*(AVG[0]-AVG[2])/AVG[0])
        intf_out.append(100*(AVG[0]-AVG[3])/AVG[0])
        intf_out.append(100*(AVG[2]-AVG[3])/AVG[2])
        return(intf_out)
    else:
        intf_out = []
        intf_out.append(100*(AVG[0]-AVG[1])/AVG[0])
        intf_out.append(100*(AVG[2]-AVG[3])/AVG[2])
        return(intf_out)


AVG=np.zeros((4,12))


for i in range(0,8):
    plate_data[i] = plate.readline().split()[1:]

for i in range(0,4):
    AVG[i]=(plate_data[i*2]+plate_data[(i*2)+1])/2

#for no sol_ant, rows 2-3 are for IDs 5-8

INTF = col_intf(AVG)
INTF = np.array(INTF)



analytics = dict()
for i in range(0,12):
    analytics[str(i)] = { "INTF" : {}}

print(AVG)

print(INTF)

def intf(m_in,ty=None):
    ty = ty #type of scorrer
    if form == 'r2':
        s = None 
        for i in range(len(thresh[0])):
            if abs(m_in) < float(thresh[0][i]):
                if s != None:
                    pass
                else:
                    s = scor_vals[0][i]
        if s == None:
            s = scor_vals[0][-1]
        s = int(s)
        return(s)
    else:
        s = None 
        for i in range(len(thresh[ty])):
            if abs(m_in) < float(thresh[ty][i]):
                if s != None:
                    pass
                else:
                    s = scor_vals[ty][i]
        if s == None:
            s = scor_vals[ty][-1]
        s = int(s)
        return(s)


def scorer():
    scores = []
    for i in range(0,12):
        if form == 'r1':
            analytics[str(i)]['INTF']['ILT3 interference'] = intf(INTF[0][i],ty=2) 
            analytics[str(i)]['INTF']['Serum interference'] = intf(INTF[1][i],ty=1)
            analytics[str(i)]['INTF']['Both interference'] = intf(INTF[2][i],ty=0)
            analytics[str(i)]['INTF']['ILT3 int (In Serum)'] = intf(INTF[3][i],ty=3)
            s = 0
            for j in analytics[str(i)]:
                for k in analytics[str(i)][j]:
                    s += analytics[str(i)][j][k]
            analytics[str(i)]['TOTAL'] = s
        else:
            analytics[str(i)]['INTF']['Serum interference r1'] = intf(INTF[0][i])
            analytics[str(i)]['INTF']['Serum interference r2'] = intf(INTF[1][i])
    '''
    if form == "r2":
        for i in range(0,12):
            s = 0
        for j in analytics[str(i)]:
            for k in analytics[str(i)][j]:
                s += analytics[str(i)][j][k]
        analytics[str(i)]['TOTAL'] = s
    '''


scorer()

print(analytics)

####finished updating up to here
def plate_2_excel(x):
    row = []
    row.append(chr(ord('@')+x+1))
    for i in plate_data[x]:
        row.append(i)
    return(row)

def avg_2_excel(x):
    if form == 'r1':
        row = []
        if x == 0:
            row.append('1%BSA/PBST')
        if x == 1:
            row.append('1%BSA/PBST + {0} ILT3'.format(con_l[-1]))
        if x == 2:
            row.append('5% HS')
        if x == 3:
            row.append('5% HS + {0} ILT3'.format(con_l[-1]))
        for i in AVG[x]:
            row.append(i)
        return(row)
    else:
        row = []
        if x == 0:
            row.append('1%BSA/PBST')
        if x == 1:
            row.append('5% HS')
        if x == 2:
            row.append('1%BSA/PBST')
        if x == 3:
            row.append('5% HS')
        for i in AVG[x]:
            row.append(i)
        return(row)

def intf_2_excel(x):
    if form == 'r1':
        row = []
        if x == 0:
            row.append('ILT3 interference')
        if x == 1:
            row.append('Serum interference')
        if x == 2:
            row.append('Both interference')
        if x == 3:
            row.append('ILT3 int (In Serum)')
        for i in INTF[x]:
            row.append(i)
        return(row)
    else:
        row = []
        if x == 0:
            row.append('Serum interference')
        if x == 1:
            row.append('Serum interference')
        for i in INTF[x]:
            row.append(i)
        return(row)

def scores_2_excel(x):
    if form == 'r1':
        row = []
        if x == 0:
            row.append('ILT3 interference')
            k = 'ILT3 interference'
        if x == 1:
            row.append('Serum interference')
            k = 'Serum interference'
        if x == 2:
            row.append('Both interference')
            k = 'Both interference'
        if x == 3:
            row.append('ILT3 interference (In Serum)')
            k = 'ILT3 int (In Serum)'
        for i in range(0,12):
            row.append(analytics[str(i)]['INTF'][k])
        return(row)
    else:
        row = []
        if x == 0:
            row.append('Serum interference')
            k = 'Serum interference r1'
            for i in range(0,12): 
                row.append(analytics[str(i)]['INTF'][k])
        else:
            row.append('Serum interference')
            k = 'Serum interference r2'
            for i in range(0,12): 
                row.append(analytics[str(i)]['INTF'][k])
        return(row)
       

exl_ary = []
if type(opts.excel_path) == str:
    data = OrderedDict()
    exl_ary.append([header[0]+' '+header[1],header[2]+' '+header[3]+' '+header[4],'','','','','','','','','','','','','','','','',h_met])
    r1 = ['',1, 2, 3, 4, 5, 6, 7, 8, 9 , 10, 11, 12]
    exl_ary.append(r1)
    for i in range(0,8):
        exl_ary.append(plate_2_excel(i))
    exl_ary.append([])
    labels = literal_eval(opts.label_path)
    label_1 = labels[0]
    label_2 = labels[1]
    label_3 = labels[2]
    label_4 = labels[3]
    exl_ary.append(['Anti-ID Label',label_1,'','',label_2,'','',label_3,'','',label_4])
    exl_ary.append(['',con_l[0],con_l[1],con_l[2],con_l[0],
                    con_l[1],con_l[2],con_l[0],con_l[1],
                    con_l[2],con_l[0],con_l[1],con_l[2]])
    if form == 'r1':
        for i in range(0,4):
            exl_ary.append(avg_2_excel(i))
        exl_ary.append([])
        exl_ary.append([])
        for i in range(0,4):
            exl_ary.append(intf_2_excel(i))
        exl_ary.append(['Anti-ID Label',label_1,'','',label_2,'','',label_3,'','',label_4])
        exl_ary.append(['Scores',con_l[0],con_l[1],con_l[2],con_l[0],
                        con_l[1],con_l[2],con_l[0],con_l[1],
                        con_l[2],con_l[0],con_l[1],con_l[2]])
        for i in range(0,4):
            exl_ary.append(scores_2_excel(i))
        row = ['total score']
        idt =[]
        for i in range(0,12):
            idt.append(analytics[str(i)]['TOTAL'])
            if (i+1)%3 == 0:
                row.append(sum(idt))
                row.append('')
                row.append('')
                idt = []
        exl_ary.append(row)
    else:
        for i in range(0,2):
            exl_ary.append(avg_2_excel(i))
        exl_ary.append([])
        exl_ary.append([])
        exl_ary.append(intf_2_excel(0))
        label_1 = labels[0]
        label_2 = labels[1]
        label_3 = labels[2]
        label_4 = labels[3]
        exl_ary.append(['Anti-ID Label',label_1,'','',label_2,'','',label_3,'','',label_4])
        exl_ary.append(['Scores',con_l[0],con_l[1],con_l[2],con_l[0],
                        con_l[1],con_l[2],con_l[0],con_l[1],
                        con_l[2],con_l[0],con_l[1],con_l[2]])
        exl_ary.append(scores_2_excel(0))
        row = ['total score']
        idt =[]
        for i in range(0,12):
            idt.append(analytics[str(i)]['INTF']['Serum interference r1'])
            if (i+1)%3 == 0:
                row.append(sum(idt))
                row.append('')
                row.append('')
                idt = []
        exl_ary.append(row)
        exl_ary.append([])
        ####
        label_1 = labels[4]
        label_2 = labels[5]
        label_3 = labels[6]
        label_4 = labels[7]
        exl_ary.append(['Anti-ID Label',label_1,'','',label_2,'','',label_3,'','',label_4])
        exl_ary.append(['',con_l[0],con_l[1],con_l[2],con_l[0],
                        con_l[1],con_l[2],con_l[0],con_l[1],
                        con_l[2],con_l[0],con_l[1],con_l[2]])
        for i in range(0,2):
            exl_ary.append(avg_2_excel(i+2))
        exl_ary.append([])
        exl_ary.append([])
        exl_ary.append(intf_2_excel(1))
        label_1 = labels[4]
        label_2 = labels[5]
        label_3 = labels[6]
        label_4 = labels[7]
        exl_ary.append(['Anti-ID Label',label_1,'','',label_2,'','',label_3,'','',label_4])
        exl_ary.append(['Scores',con_l[0],con_l[1],con_l[2],con_l[0],
                        con_l[1],con_l[2],con_l[0],con_l[1],
                        con_l[2],con_l[0],con_l[1],con_l[2]])
        exl_ary.append(scores_2_excel(1))
        row = ['total score']
        idt =[]
        for i in range(0,12):
            idt.append(analytics[str(i)]['INTF']['Serum interference r2'])
            if (i+1)%3 == 0:
                row.append(sum(idt))
                row.append('')
                row.append('')
                idt = []
        exl_ary.append(row)
        print(exl_ary)
        
    data.update({"Sheet 1": exl_ary})
    save_data(opts.excel_path, data)


#workbook = load_workbook(filename="data.xlsx")
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill, Fill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
workbook = load_workbook(opts.excel_path)
sheet = workbook.active

t =  Side(border_style="thin")

top = Border(top=t,
                        right=t,
                        bottom=None,
                        left=t)

bottom = Border(top=None,
                        right=t,
                        bottom=t,
                        left=t)

left = Border(top=None,
                        right=None,
                        bottom=None,
                        left=t)

right = Border(top=None,
                        right=t,
                        bottom=None,
                        left=None)

sheet.merge_cells('B12:D12')
sheet.merge_cells('E12:G12')
sheet.merge_cells('H12:J12')
sheet.merge_cells('K12:M12')
sheet["A12"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
sheet["B12"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
sheet["E12"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
sheet["H12"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
sheet["K12"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")

if form == 'r1':
    for i in range(13,30):
        sheet["B{0}".format(i)].border = left
        sheet["D{0}".format(i)].border = right
        sheet["G{0}".format(i)].border = right
        sheet["J{0}".format(i)].border = right
        sheet["M{0}".format(i)].border = right

    sheet["B12"].alignment = Alignment(horizontal="center")
    sheet["E12"].alignment = Alignment(horizontal="center")
    sheet["H12"].alignment = Alignment(horizontal="center")
    sheet["K12"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}12".format(chr(ord('@')+i))].border = top

    sheet.merge_cells('B24:D24')
    sheet.merge_cells('E24:G24')
    sheet.merge_cells('H24:J24')
    sheet.merge_cells('K24:M24')
    sheet["A24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["B24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["E24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["H24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["K24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")

    sheet["B24"].alignment = Alignment(horizontal="center")
    sheet["E24"].alignment = Alignment(horizontal="center")
    sheet["H24"].alignment = Alignment(horizontal="center")
    sheet["K24"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}24".format(chr(ord('@')+i))].border = top


    sheet.merge_cells('B30:D30')
    sheet.merge_cells('E30:G30')
    sheet.merge_cells('H30:J30')
    sheet.merge_cells('K30:M30')
    sheet["B30"].alignment = Alignment(horizontal="center")
    sheet["E30"].alignment = Alignment(horizontal="center")
    sheet["H30"].alignment = Alignment(horizontal="center")
    sheet["K30"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}30".format(chr(ord('@')+i))].border = bottom

    sheet["A15"].alignment = Alignment(wrapText=True)
    sheet["A25"].alignment = Alignment(wrapText=True)
    sheet["A29"].alignment = Alignment(wrapText=True)

    sheet.column_dimensions["A"].width = 18.5703125
    sheet.row_dimensions[15].height = 30.0
    sheet.row_dimensions[25].height = 45.0
    sheet.row_dimensions[29].height = 30
else:
    for i in range(13,34):
        sheet["B{0}".format(i)].border = left
        sheet["D{0}".format(i)].border = right
        sheet["G{0}".format(i)].border = right
        sheet["J{0}".format(i)].border = right
        sheet["M{0}".format(i)].border = right

    sheet["B12"].alignment = Alignment(horizontal="center")
    sheet["E12"].alignment = Alignment(horizontal="center")
    sheet["H12"].alignment = Alignment(horizontal="center")
    sheet["K12"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}12".format(chr(ord('@')+i))].border = top

    sheet.merge_cells('B19:D19')
    sheet.merge_cells('E19:G19')
    sheet.merge_cells('H19:J19')
    sheet.merge_cells('K19:M19')
    sheet["A19"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["B19"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["E19"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["H19"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["K19"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")

    sheet["B19"].alignment = Alignment(horizontal="center")
    sheet["E19"].alignment = Alignment(horizontal="center")
    sheet["H19"].alignment = Alignment(horizontal="center")
    sheet["K19"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}19".format(chr(ord('@')+i))].border = top


    sheet.merge_cells('B22:D22')
    sheet.merge_cells('E22:G22')
    sheet.merge_cells('H22:J22')
    sheet.merge_cells('K22:M22')
    sheet["B22"].alignment = Alignment(horizontal="center")
    sheet["E22"].alignment = Alignment(horizontal="center")
    sheet["H22"].alignment = Alignment(horizontal="center")
    sheet["K22"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}22".format(chr(ord('@')+i))].border = bottom

    #sheet["A15"].alignment = Alignment(wrapText=True)
    #sheet["A25"].alignment = Alignment(wrapText=True)
    #sheet["A29"].alignment = Alignment(wrapText=True)

    sheet.column_dimensions["A"].width = 18.5703125
    #sheet.row_dimensions[15].height = 30.0
    sheet.row_dimensions[20].height = 45.0 #scores
    #sheet.row_dimensions[29].height = 30
    ###
    sheet.merge_cells('B24:D24')
    sheet.merge_cells('E24:G24')
    sheet.merge_cells('H24:J24')
    sheet.merge_cells('K24:M24')
    sheet["B24"].alignment = Alignment(horizontal="center")
    sheet["E24"].alignment = Alignment(horizontal="center")
    sheet["H24"].alignment = Alignment(horizontal="center")
    sheet["K24"].alignment = Alignment(horizontal="center")
    
    sheet["A24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["B24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["E24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["H24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["K24"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    for i in range(2,14):
        sheet["{0}24".format(chr(ord('@')+i))].border = top

    sheet.merge_cells('B31:D31')
    sheet.merge_cells('E31:G31')
    sheet.merge_cells('H31:J31')
    sheet.merge_cells('K31:M31')
    sheet["A31"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["B31"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["E31"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["H31"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")
    sheet["K31"].fill = PatternFill(patternType='solid', fgColor="FFFFFF00")

    sheet["B31"].alignment = Alignment(horizontal="center")
    sheet["E31"].alignment = Alignment(horizontal="center")
    sheet["H31"].alignment = Alignment(horizontal="center")
    sheet["K31"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}31".format(chr(ord('@')+i))].border = top


    sheet.merge_cells('B34:D34')
    sheet.merge_cells('E34:G34')
    sheet.merge_cells('H34:J34')
    sheet.merge_cells('K34:M34')
    sheet["B34"].alignment = Alignment(horizontal="center")
    sheet["E34"].alignment = Alignment(horizontal="center")
    sheet["H34"].alignment = Alignment(horizontal="center")
    sheet["K34"].alignment = Alignment(horizontal="center")
    for i in range(2,14):
        sheet["{0}34".format(chr(ord('@')+i))].border = bottom

    #sheet["A15"].alignment = Alignment(wrapText=True)
    #sheet["A25"].alignment = Alignment(wrapText=True)
    #sheet["A29"].alignment = Alignment(wrapText=True)

    sheet.column_dimensions["A"].width = 18.5703125
    #sheet.row_dimensions[15].height = 30.0
    sheet.row_dimensions[34].height = 45.0
    #sheet.row_dimensions[29].height = 30
workbook.save(opts.excel_path)
print("done")


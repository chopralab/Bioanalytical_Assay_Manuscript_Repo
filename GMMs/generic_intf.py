import numpy as np
from optparse import OptionParser
from pyexcel_xlsx import save_data
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill, Fill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

import pandas as pd
import random
seed = 7
np.random.seed(seed)
random.seed(7)

excel_path = "Compiled MK-0482 part 1 data for IT 12-4-18.xlsx"

workbook = load_workbook(excel_path)
sheet = workbook.active

for i in range(30,40):
    sheet["B{0}".format(i)].alignment = Alignment(wrapText=True)


rinf_dict = {}
set_a = ['1','2','3']
set_b = ['b','e','h','k','n','q','t','w','z','ac','af',
         'ai','al','ao','ar','au','ax','ba']
ak = 0
for i in set_a:
    print(ak)
    rinf_dict[i] = {}
    for j in set_b:
        rinf_dict[i][sheet[j+"1"].value.split()[0]] = {'25 ug/mL':{},'2.5 ug/mL':{},'0.25 ug/mL':{}}
        for j2 in range(29+ak,33+ak):
            for i2,k2 in zip(sheet['{0}2:bi2'.format(j)][0][0:3], sheet['{0}{1}:bi{1}'.format(j,j2)][0][0:3]):
                rinf_dict[i][sheet[j+"1"].value.split()[0]][i2.value][sheet['a{0}'.format(j2)].value]=k2.value
    ak+=20

rinf_list = []
ak = 0
for i in set_a:
    print(ak)
    rinf_list.append([])
    for j in set_b:
        
        for j2 in range(29+ak,33+ak):
            for i2,k2 in zip(sheet['{0}2:bi2'.format(j)][0][0:3], sheet['{0}{1}:bi{1}'.format(j,j2)][0][0:3]):
                rinf_dict[i][sheet[j+"1"].value.split()[0]][i2.value][sheet['a{0}'.format(j2)].value]=k2.value
    ak+=20




rinf_1 = rinf_dict['1']
rinf_1_BSAPBST = []
for i in rinf_1:
    for i2 in rinf_1[i]:
        print(i2)
        rinf_1_BSAPBST.append(rinf_1[i][i2]['ILT3 interference'])



rinf_1_BSAPBST_25 = []
rinf_1_BSAPBST_2p5 = []
rinf_1_BSAPBST_p25 = []
c = 1
for i in rinf_1_BSAPBST:
    if c%2==0:
        rinf_1_BSAPBST_2p5.append(i)
    elif c%3==0:
        rinf_1_BSAPBST_p25.append(i)
        c=0
    else:
        rinf_1_BSAPBST_25.append(i)
    c+=1



rinf_1_BSAPBSTILT3 = []
for i in rinf_1:
    for i2 in rinf_1[i]:
        rinf_1_BSAPBSTILT3.append(rinf_1[i][i2]['Serum interference'])



rinf_1_BSAPBSTILT3_25 = []
rinf_1_BSAPBSTILT3_2p5 = []
rinf_1_BSAPBSTILT3_p25 = []
c = 1
for i in rinf_1_BSAPBSTILT3:
    if c%2==0:
        rinf_1_BSAPBSTILT3_2p5.append(i)
    elif c%3==0:
        rinf_1_BSAPBSTILT3_p25.append(i)
        c=0
    else:
        rinf_1_BSAPBSTILT3_25.append(i)
    c+=1



rinf_1_HS = []
for i in rinf_1:
    for i2 in rinf_1[i]:
        rinf_1_HS.append(rinf_1[i][i2]['Both interference'])



rinf_1_HS_25 = []
rinf_1_HS_2p5 = []
rinf_1_HS_p25 = []
c = 1
for i in rinf_1_HS:
    if c%2==0:
        rinf_1_HS_2p5.append(i)
    elif c%3==0:
        rinf_1_HS_p25.append(i)
        c=0
    else:
        rinf_1_HS_25.append(i)
    c+=1
    
    

rinf_1_HS25 = []
for i in rinf_1:
    for i2 in rinf_1[i]:
        rinf_1_HS25.append(rinf_1[i][i2]['ILT3 interference (In Serum)'])



rinf_1_HS25_25 = []
rinf_1_HS25_2p5 = []
rinf_1_HS25_p25 = []
c = 1
for i in rinf_1_HS25:
    if c%2==0:
        rinf_1_HS25_2p5.append(i)
    elif c%3==0:
        rinf_1_HS25_p25.append(i)
        c=0
    else:
        rinf_1_HS25_25.append(i)
    c+=1    



###########
rinf_2 = rinf_dict['2']
rinf_2_BSAPBST = []
for i in rinf_2:
    for i2 in rinf_2[i]:
        rinf_2_BSAPBST.append(rinf_2[i][i2]['ILT3 interference'])



rinf_2_BSAPBST_25 = []
rinf_2_BSAPBST_2p5 = []
rinf_2_BSAPBST_p25 = []
c = 1
for i in rinf_2_BSAPBST:
    if c%2==0:
        rinf_2_BSAPBST_2p5.append(i)
    elif c%3==0:
        rinf_2_BSAPBST_p25.append(i)
        c=0
    else:
        rinf_2_BSAPBST_25.append(i)
    c+=1



rinf_2_BSAPBSTILT3 = []
for i in rinf_2:
    for i2 in rinf_2[i]:
        rinf_2_BSAPBSTILT3.append(rinf_2[i][i2]['Serum interference'])



rinf_2_BSAPBSTILT3_25 = []
rinf_2_BSAPBSTILT3_2p5 = []
rinf_2_BSAPBSTILT3_p25 = []
c = 1
for i in rinf_2_BSAPBSTILT3:
    if c%2==0:
        rinf_2_BSAPBSTILT3_2p5.append(i)
    elif c%3==0:
        rinf_2_BSAPBSTILT3_p25.append(i)
        c=0
    else:
        rinf_2_BSAPBSTILT3_25.append(i)
    c+=1



rinf_2_HS = []
for i in rinf_2:
    for i2 in rinf_2[i]:
        rinf_2_HS.append(rinf_2[i][i2]['Both interference'])



rinf_2_HS_25 = []
rinf_2_HS_2p5 = []
rinf_2_HS_p25 = []
c = 1
for i in rinf_2_HS:
    if c%2==0:
        rinf_2_HS_2p5.append(i)
    elif c%3==0:
        rinf_2_HS_p25.append(i)
        c=0
    else:
        rinf_2_HS_25.append(i)
    c+=1
    

    
rinf_2_HS25 = []
for i in rinf_2:
    for i2 in rinf_2[i]:
        rinf_2_HS25.append(rinf_2[i][i2]['ILT3 interference (In Serum)'])



rinf_2_HS25_25 = []
rinf_2_HS25_2p5 = []
rinf_2_HS25_p25 = []
c = 1
for i in rinf_2_HS25:
    if c%2==0:
        rinf_2_HS25_2p5.append(i)
    elif c%3==0:
        rinf_2_HS25_p25.append(i)
        c=0
    else:
        rinf_2_HS25_25.append(i)
    c+=1    


####
rinf_3 = rinf_dict['3']
rinf_3_BSAPBST = []
for i in rinf_3:
    for i2 in rinf_3[i]:
        rinf_3_BSAPBST.append(rinf_3[i][i2]['ILT3 interference'])



rinf_3_BSAPBST_25 = []
rinf_3_BSAPBST_2p5 = []
rinf_3_BSAPBST_p25 = []
c = 1
for i in rinf_3_BSAPBST:
    if c%2==0:
        rinf_3_BSAPBST_2p5.append(i)
    elif c%3==0:
        rinf_3_BSAPBST_p25.append(i)
        c=0
    else:
        rinf_3_BSAPBST_25.append(i)
    c+=1



rinf_3_BSAPBSTILT3 = []
for i in rinf_3:
    for i2 in rinf_3[i]:
        rinf_3_BSAPBSTILT3.append(rinf_3[i][i2]['Serum interference'])



rinf_3_BSAPBSTILT3_25 = []
rinf_3_BSAPBSTILT3_2p5 = []
rinf_3_BSAPBSTILT3_p25 = []
c = 1
for i in rinf_3_BSAPBSTILT3:
    if c%2==0:
        rinf_3_BSAPBSTILT3_2p5.append(i)
    elif c%3==0:
        rinf_3_BSAPBSTILT3_p25.append(i)
        c=0
    else:
        rinf_3_BSAPBSTILT3_25.append(i)
    c+=1



rinf_3_HS = []
for i in rinf_3:
    for i2 in rinf_3[i]:
        rinf_3_HS.append(rinf_3[i][i2]['Both interference'])



rinf_3_HS_25 = []
rinf_3_HS_2p5 = []
rinf_3_HS_p25 = []
c = 1
for i in rinf_3_HS:
    if c%2==0:
        rinf_3_HS_2p5.append(i)
    elif c%3==0:
        rinf_3_HS_p25.append(i)
        c=0
    else:
        rinf_3_HS_25.append(i)
    c+=1
    
    

rinf_3_HS25 = []
for i in rinf_3:
    for i2 in rinf_3[i]:
        rinf_3_HS25.append(rinf_3[i][i2]['ILT3 interference (In Serum)'])



rinf_3_HS25_25 = []
rinf_3_HS25_2p5 = []
rinf_3_HS25_p25 = []
c = 1
for i in rinf_3_HS25:
    if c%2==0:
        rinf_3_HS25_2p5.append(i)
    elif c%3==0:
        rinf_3_HS25_p25.append(i)
        c=0
    else:
        rinf_3_HS25_25.append(i)
    c+=1    


#each list has the 3 concentrations 
out =[[rinf_1_BSAPBST, rinf_1_BSAPBSTILT3, rinf_1_HS, rinf_1_HS25],[rinf_2_BSAPBST,rinf_2_BSAPBSTILT3, rinf_2_HS, rinf_2_HS25],[rinf_3_BSAPBST, rinf_3_BSAPBSTILT3, rinf_3_HS, rinf_3_HS25]]
out2 = []
for replicate in out:
    out2.append([])
    for experiment in replicate:
        out2[-1].append([[],[],[]])
        c = 0
        print(experiment)
        print(out2)
        print(out2[-1])
        print(out2[-1][-1])
        print(out2[-1][-1][c])
        for concentration in experiment:
            out2[-1][-1][c].append(concentration)
            c += 1
            if c == 3:
                c = 0

np.save("reverse_data",out2)


import sklearn
from sklearn import metrics
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from sklearn.decomposition import PCA
import copy
import numpy as np
import itertools
from scipy import linalg
import matplotlib.pyplot as plt
import matplotlib as mpl
from sklearn import mixture
from scipy import linalg
from mpl_toolkits.mplot3d import Axes3D
import umap

def pk_bd(inp):
    org = []
    org.append(min(inp[:,0]))
    org.append(min(inp[:,1]))
    org.append(min(inp[:,2]))
    org.append(min(inp[:,3]))
    org.append(min(inp[:,4]))
    org.append(min(inp[:,5]))
    org.append(min(inp[:,6]))
    org.append(min(inp[:,7]))
    org.append(min(inp[:,8]))
    org.append(min(inp[:,9]))
    org.append(min(inp[:,10]))
    org.append(min(inp[:,11]))
    dist =[]
    for i in inp:
        dist.append(np.linalg.norm(i - org))
    return(dist)

mmft = MinMaxScaler().fit_transform
metric="aic"
replicate_names = ["1xA","1xB","40x"]
condition_names = ["Soluble Antigen Interference","Matrix Interference","Combination Interference","Soluble Antigen in Matrix Interference"]
concentration_names = ["100x","10x","1x"]
'''
p = 0
r = 0 #replicate names
for replicate in out2:
    e = 0 #CONDITION_NAMES
    gmm_scores = [["Anti-ID & Concentration","Soluble Antigen Interference","Matrix Interference","Combination Interference","Soluble Antigen in Matrix Interference", "Total"]]
    for conc_c in range(3):
        for ID in range(18):
            gmm_scores.append([])
            gmm_scores[ID+1+(conc_c*18)].append("Anti-ID {0} [{1}]".format(ID+1,concentration_names[conc_c]))
    gmm_bic = [["Analytical Metric & Condition","4 Components",
                "5 Components", "6 Components", "7 Components", "8 Components", "9 Components",
                "10 Components"]]
    for experiment in replicate:
        c = 0
        for concentration in experiment:
            gmm_bic.append([])
            gmm_bic[-1].append('{0} with Anti-ID [{1}]'.format(condition_names[e],concentration_names[c]))
            inp = np.abs(concentration).flatten()
            avv2 = np.array([inp,np.zeros(len(inp))]).transpose()
            lowest_bic = np.infty
            bic = []
            bil = []
            n_components_range = range(4, 11)
            cv_types = ['diag']
            for cv_type in cv_types:
                for n_components in n_components_range:
                    # Fit a Gaussian mixture with EM
                    gmm = mixture.GaussianMixture(n_components=n_components,
                                                  covariance_type=cv_type)
                    gmm.fit(avv2)
                    if metric == 'bic':
                        bic.append(gmm.bic(avv2))
                    elif metric == 'aic':
                        bic.append(gmm.aic(avv2))
                    elif metric == 'score':
                        bic.append(-1*gmm.score(avv2))
                    #bic.append(gmm.aic(avv2)-(gmm.score(avv2)))
                    gmm_bic[-1].append(bic[-1]) 
                    if bic[-1] < lowest_bic:
                        lowest_bic = bic[-1]
                        best_gmm = gmm    
            bic = np.array(bic)
            color_iter = itertools.cycle(['navy', 'turquoise', 'cornflowerblue',
                                        'darkorange','maroon','gold','thistle',
                                        'peachpuff','lightsteelblue','pink'])
            clf = best_gmm
            bars = []
            # Plot the BIC scores
            plt.figure(figsize=(6,2))
            # Plot the winner
            splot = plt.subplot(1, 1, 1)
            Y_ = clf.predict(avv2)
            #pca = PCA(n_components=2)
            #pc = pca.fit_transform(ssft(avv))
            print(avv2[0])
            print(avv2[Y_ == i])
            avv = inp
            plt.scatter(min(avv),0,c='k',s=30)
            color_iter = itertools.cycle(['#EAA568', '#EAD24A', '#CFEA4A', '#8AEA4A', '#4AEA6A', '#4AEACD', '#4AE2EA', '#4AA7EA', '#4A62EA', '#614AEA'])
            colors = ['#EAA568', '#EAD24A', '#CFEA4A', '#8AEA4A', '#4AEA6A', '#4AEACD', '#4AE2EA', '#4AA7EA', '#4A62EA', '#614AEA']
            Y_v_list = []
            for i in range(max(Y_)+1):
                Y_v = [avv[v[0]] for v in enumerate(Y_) if v[1] == i]
                Y_v_list.append(np.mean(Y_v))
            cl = np.arange(10)
            cl = list(cl)
            for sort_val in enumerate(np.argsort(Y_v_list)):
                sc = colors[sort_val[0]]
                print(sc)
                cl[sort_val[1]] = str(sc)

            Y_i = np.argsort(Y_v_list)
            Y_i = list(Y_i)
            Y_i.reverse()
            for index_v in enumerate(Y_):
                gmm_scores[index_v[0]+1+(18*c)].append(Y_i.index(index_v[1]))

            for i, (mean, cov, color) in enumerate(zip(clf.means_, clf.covariances_,
                                                    color_iter)):
                if not np.any(Y_ == i):
                    continue
                plt.scatter(avv2[Y_ == i,0], avv2[Y_ == i, 1], 10, color=cl[i])
            nc = np.mod(bic.argmin(), len(n_components_range)) + 2
            ct = cv_types[int(np.floor(bic.argmin() / len(n_components_range)))]
            plt.plot([20,20],[-0.1,.1],c='grey')
            plt.plot([10,10],[-0.1,.1],c='grey')
            splot.set_ylim([-.01, .01])
            #plt.xticks(())
            if max(avv) < 100:
                p_end = 30
                if max(avv) > 25:
                    p_end = (round(max(avv))//5)*5+10
                plt.xticks(np.arange(0,p_end,5))
            else:
                p_end = (round(max(avv))//10)*10+20
                plt.xticks(np.arange(0,p_end,10))
            plt.xlabel(condition_names[e])
            plt.yticks(())
            plt.title('Antigen {0}, Anti-ID Concentration: {1}'.format(replicate_names[r],concentration_names[c]))
            plt.subplots_adjust(hspace=.35, bottom=.02)
            plt.tight_layout()
            plt.savefig("plots/reverse_gmm_{0}.svg".format(p),dpi=300)
            plt.savefig("plots/reverse_gmm_{0}".format(p),dpi=300)
            plt.show()
            c += 1
            p += 1
        e += 1
    gmm_bic = pd.DataFrame(gmm_bic)
    gmm_bic.to_csv("gmm_reverse_aic_{0}.csv".format(replicate_names[r]),index=False,header=False)
    c2 = 1
    for gmm_i in gmm_scores[1:]:
        gmm_scores[c2].append(sum(gmm_scores[c2][1:]))
        c2 += 1
    gmm_scores = pd.DataFrame(gmm_scores)
    gmm_scores.to_csv("gmm_reverse_scores_{0}.csv".format(replicate_names[r]),index=False,header=False)
    r += 1
'''
#### all 


p = 0
r = 0 #replicate names
for replicate in out2:
    e = 0 #CONDITION_NAMES
    gmm_scores = [["Anti-ID","Score"]]
    for ID in range(18):
        gmm_scores.append([f"Anti-ID {ID+1}"])
    gmm_bic = [["Components","4 Components",
                "5 Components", "6 Components", "7 Components", "8 Components", "9 Components",
                "10 Components"]]
    replicate = np.array(replicate)
    inp = np.abs(replicate.reshape((12,18)))
    inp = inp.transpose()
    scaler_m = preprocessing.MinMaxScaler()
    scaler_m.fit(inp)
    avv2 = scaler_m.transform(inp)
    gmm_bic.append(["Scores"])
    lowest_bic = np.infty
    bic = []
    bil = []
    n_components_range = range(4, 11)
    cv_types = ['full']
    for cv_type in cv_types:
        for n_components in n_components_range:
            # Fit a Gaussian mixture with EM
            gmm = mixture.GaussianMixture(n_components=n_components,
                                            covariance_type=cv_type)
            gmm.fit(avv2)
            if metric == 'bic':
                bic.append(gmm.bic(avv2))
            elif metric == 'aic':
                bic.append(gmm.aic(avv2))
            elif metric == 'score':
                bic.append(-1*gmm.score(avv2))
            #bic.append(gmm.aic(avv2)-(gmm.score(avv2)))
            gmm_bic[-1].append(bic[-1]) 
            if bic[-1] < lowest_bic:
                lowest_bic = bic[-1]
                best_gmm = gmm    
    bic = np.array(bic)
    clf = best_gmm
    # Plot the BIC scores
    plt.figure()
    # Plot the winner
    splot = plt.subplot(1, 1, 1)
    Y_ = clf.predict(avv2)
    #pca = PCA(n_components=2)
    #pc = pca.fit_transform(ssft(avv))
    reducer = umap.UMAP(n_components=2,min_dist=0.25,random_state=2)
    reducer.fit(mmft(avv2))#used to have standardizer here
    pc = reducer.transform(mmft(avv2))
    refrence = reducer.transform([[0,0,0,0,0,0,0,0,0,0,0,0]])

    print(avv2[0])
    print(avv2[Y_ == i])
    plt.scatter(refrence[0,0],refrence[0,1],c='k',s=70)
    
    color_iter = itertools.cycle(['#1100FF', '#4AA7EA','#4AE2EA','#4AEACD','#8AEA4A' ,'#CFEA4A','#EAD24A','#EAA568', "#EA644A" ,"#E33535"])
    colors = ['#1100FF', '#4AA7EA','#4AE2EA','#4AEACD','#8AEA4A' ,'#CFEA4A','#EAD24A','#EAA568', "#EA644A" ,"#E33535"]
    
    fnd = pk_bd(avv2)
    
    scord = np.argmin(fnd)
    plt.scatter(pc[scord,0],pc[scord,1],c='k',s=70)
    LLL = ['#1100FF', '#4AA7EA','#4AE2EA','#4AEACD','#8AEA4A' ,'#CFEA4A','#EAD24A','#EAA568', "#EA644A" ,"#E33535"]
    LLL.reverse()
    color_iter = itertools.cycle(LLL)
    Y_v_list = []
    for i in range(max(Y_)+1):
        Y_v = [fnd[v[0]] for v in enumerate(Y_) if v[1] == i]
        Y_v_list.append(np.mean(Y_v))#appends max value by increasing Y_ labels
    Y_v_list_2 = copy.deepcopy(Y_v_list)
    print(Y_v_list)
    colors_p = copy.deepcopy(LLL)
    cl = np.arange(10)
    cl = list(cl)
    Y_score = []
    for sort_val in enumerate(np.argsort(Y_v_list)):
        print(cl[sort_val[1]])
        Y_score.append(cl[sort_val[1]])
        sc = colors_p[sort_val[0]]
        print(sc)
        cl[sort_val[1]] = str(sc)

    #mapping Y_ indexes to scores
    Y_score.reverse()
    
    for index_v in enumerate(Y_):
        gmm_scores[index_v[0]+1].append(Y_score.index(index_v[1]))

    for i, (mean, cov, color) in enumerate(zip(clf.means_, clf.covariances_,
                                            color_iter)):
        if not np.any(Y_ == i):
            continue
        plt.scatter(pc[Y_ == i,0], pc[Y_ == i, 1], 40, color=cl[i])
    nc = np.mod(bic.argmin(), len(n_components_range)) + 2
    ct = cv_types[int(np.floor(bic.argmin() / len(n_components_range)))]
    #plt.plot([20,20],[-0.1,.1],c='grey')
    #plt.plot([10,10],[-0.1,.1],c='grey')
    #plt.xticks(())
    plt.xlabel('Component_1',fontsize=12)
    plt.ylabel('Component_2',fontsize=12)
    plt.title('UMAP of GMM Clustered Generic Screenings\n({0})'.format(replicate_names[r]))
    plt.subplots_adjust(hspace=.35, bottom=.02)
    plt.tight_layout()
    plt.savefig(f"plots/reverse_gmm_full_{replicate_names[r]}.pdf",dpi=300)
    plt.savefig(f"plots/reverse_gmm_full_{replicate_names[r]}",dpi=300)
    plt.xticks(fontsize=12)
    plt.show()
    gmm_bic = pd.DataFrame(gmm_bic)
    gmm_bic.to_csv("gmm_reverse_full_aic_{0}.csv".format(replicate_names[r]),index=False,header=False)
    c2 = 1
    for gmm_i in gmm_scores[1:]:
        gmm_scores[c2].append(sum(gmm_scores[c2][1:]))
        c2 += 1
    gmm_scores = pd.DataFrame(gmm_scores)
    gmm_scores.to_csv("gmm_reverse_scores_full_{0}.csv".format(replicate_names[r]),index=False,header=False)
    r += 1